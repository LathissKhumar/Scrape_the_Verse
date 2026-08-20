"""
SEO Exporter - Exports detailed audit results to JSON and formatted Excel (.xlsx) workbooks.
Uses openpyxl to generate styled, multi-sheet workbooks with professional formatting.
"""

import json
import os
from typing import Dict, Any, List

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def export_to_json(state: Dict[str, Any], filepath: str) -> None:
    """Save complete structured audit state to JSON file."""
    os.makedirs(os.path.dirname(os.path.abspath(filepath)) or '.', exist_ok=True)
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(state, f, indent=2, default=str)


def export_to_excel(state: Dict[str, Any], filepath: str) -> bool:
    """
    Export detailed SEO audit into a clean, multi-tab Excel workbook (.xlsx).
    Sheets included:
      1. Audit Overview & Scores
      2. Priority Action Items
      3. Detected Issues (with Evidence)
      4. Crawled Pages (full metadata)
      5. Link Graph
      6. Performance & PageSpeed
    """
    if not OPENPYXL_AVAILABLE:
        print("Warning: openpyxl is not installed. Run `pip install openpyxl` to export to Excel.")
        return False

    os.makedirs(os.path.dirname(os.path.abspath(filepath)) or '.', exist_ok=True)
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Styling Palettes
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    title_font = Font(name="Calibri", size=14, bold=True, color="0F172A")
    subtitle_font = Font(name="Calibri", size=10, italic=True, color="64748B")
    
    bold_font = Font(name="Calibri", size=11, bold=True)
    regular_font = Font(name="Calibri", size=11)
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    fill_pass = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")   # Green
    fill_warn = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")   # Yellow
    fill_fail = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")   # Red
    fill_info = PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid")   # Blue

    # -------------------------------------------------------------------------
    # Sheet 1: Audit Overview
    # -------------------------------------------------------------------------
    ws_summary = wb.create_sheet(title="Overview & Scores")
    ws_summary.views.sheetView[0].showGridLines = True

    ws_summary.cell(row=1, column=1, value="SEO Audit Executive Overview").font = title_font
    ws_summary.cell(row=2, column=1, value=f"Target URL: {state.get('url', '')}").font = subtitle_font

    summary_data = [
        ("Overall SEO Health Score", f"{state.get('overall_seo_score', 0)} / 100"),
        ("Pages Crawled", len(state.get('pages', []))),
        ("Links Analyzed", len(state.get('links', []))),
        ("Total Issues Detected", len(state.get('issues', []))),
        ("Crawl Duration", f"{state.get('crawl_summary', {}).get('duration_seconds', 0)} seconds"),
        ("Audit Status", state.get('status', 'completed').upper()),
    ]

    ws_summary.cell(row=4, column=1, value="Metric").font = header_font
    ws_summary.cell(row=4, column=1).fill = header_fill
    ws_summary.cell(row=4, column=2, value="Value").font = header_font
    ws_summary.cell(row=4, column=2).fill = header_fill

    for i, (k, v) in enumerate(summary_data, start=5):
        c1 = ws_summary.cell(row=i, column=1, value=k)
        c2 = ws_summary.cell(row=i, column=2, value=v)
        c1.font = bold_font
        c2.font = regular_font
        c1.border = thin_border
        c2.border = thin_border

    # Category Scores Table
    ws_summary.cell(row=12, column=1, value="Category Score Breakdown").font = Font(name="Calibri", size=12, bold=True)
    cat_headers = ["Category", "Score", "Status", "Summary"]
    for col_idx, h in enumerate(cat_headers, start=1):
        cell = ws_summary.cell(row=14, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center" if col_idx in (2, 3) else "left")

    categories_info = [
        ("Technical SEO", state.get("technical_audit", {})),
        ("On-Page SEO", state.get("onpage_audit", {})),
        ("Content Quality", state.get("content_audit", {})),
        ("Performance", state.get("performance_audit", {})),
        ("Structured Data", state.get("schema_audit", {})),
        ("Local SEO", state.get("local_audit", {})),
    ]

    for i, (name, audit) in enumerate(categories_info, start=15):
        score = audit.get("score", 100)
        status = audit.get("status", "passed").upper()
        summary = audit.get("summary", "")

        c1 = ws_summary.cell(row=i, column=1, value=name)
        c2 = ws_summary.cell(row=i, column=2, value=f"{score}/100")
        c3 = ws_summary.cell(row=i, column=3, value=status)
        c4 = ws_summary.cell(row=i, column=4, value=summary)

        for c in (c1, c2, c3, c4):
            c.font = regular_font
            c.border = thin_border

        c2.alignment = Alignment(horizontal="center")
        c3.alignment = Alignment(horizontal="center")

        if status == "PASSED":
            c3.fill = fill_pass
        elif status == "WARNING":
            c3.fill = fill_warn
        else:
            c3.fill = fill_fail

    # -------------------------------------------------------------------------
    # Sheet 2: Priority Action Items
    # -------------------------------------------------------------------------
    ws_actions = wb.create_sheet(title="Priority Action Items")
    ws_actions.views.sheetView[0].showGridLines = True
    ws_actions.freeze_panes = "A2"

    action_headers = [
        "Priority", "Category", "Issue Title", "Recommended Action",
        "Impact Score (1-10)", "Estimated Effort", "Affected Pages Count"
    ]
    for col_idx, h in enumerate(action_headers, start=1):
        cell = ws_actions.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center" if col_idx in (1, 5, 6, 7) else "left")

    actions = state.get("priority_action_items", [])
    for row_idx, item in enumerate(actions, start=2):
        prio = item.get("priority", 3)
        ws_actions.cell(row=row_idx, column=1, value=f"Priority {prio}").alignment = Alignment(horizontal="center")
        ws_actions.cell(row=row_idx, column=2, value=item.get("category", ""))
        ws_actions.cell(row=row_idx, column=3, value=item.get("title", ""))
        ws_actions.cell(row=row_idx, column=4, value=item.get("action", ""))
        ws_actions.cell(row=row_idx, column=5, value=item.get("impact_score", 5)).alignment = Alignment(horizontal="center")
        ws_actions.cell(row=row_idx, column=6, value=item.get("estimated_effort", "medium").title()).alignment = Alignment(horizontal="center")
        ws_actions.cell(row=row_idx, column=7, value=item.get("affected_count", 0)).alignment = Alignment(horizontal="center")

        # Color-code priority
        p_cell = ws_actions.cell(row=row_idx, column=1)
        if prio == 1:
            p_cell.fill = fill_fail
        elif prio == 2:
            p_cell.fill = fill_warn
        elif prio == 3:
            p_cell.fill = fill_info
        else:
            p_cell.fill = fill_pass

        for c_idx in range(1, 8):
            cell = ws_actions.cell(row=row_idx, column=c_idx)
            cell.font = regular_font
            cell.border = thin_border

    # -------------------------------------------------------------------------
    # Sheet 3: Detected Issues (with Evidence)
    # -------------------------------------------------------------------------
    ws_issues = wb.create_sheet(title="Detected Issues")
    ws_issues.views.sheetView[0].showGridLines = True
    ws_issues.freeze_panes = "A2"

    issue_headers = ["URL", "Issue Type", "Category", "Severity", "Issue Name", "Details", "Evidence (JSON/Data)"]
    for col_idx, h in enumerate(issue_headers, start=1):
        cell = ws_issues.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center" if col_idx in (2, 3, 4) else "left")

    issues_list = state.get("issues", [])
    for row_idx, iss in enumerate(issues_list, start=2):
        sev = (iss.get("severity") or "medium").lower()
        ws_issues.cell(row=row_idx, column=1, value=iss.get("url", ""))
        ws_issues.cell(row=row_idx, column=2, value=iss.get("type", "")).alignment = Alignment(horizontal="center")
        ws_issues.cell(row=row_idx, column=3, value=iss.get("category", "")).alignment = Alignment(horizontal="center")
        
        sev_cell = ws_issues.cell(row=row_idx, column=4, value=sev.upper())
        sev_cell.alignment = Alignment(horizontal="center")
        if sev in ("critical", "error"):
            sev_cell.fill = fill_fail
        elif sev in ("high", "warning"):
            sev_cell.fill = fill_warn
        elif sev == "info":
            sev_cell.fill = fill_info

        ws_issues.cell(row=row_idx, column=5, value=iss.get("issue", ""))
        ws_issues.cell(row=row_idx, column=6, value=iss.get("details", ""))
        
        evidence_str = json.dumps(iss.get("evidence", {}), default=str) if iss.get("evidence") else ""
        ws_issues.cell(row=row_idx, column=7, value=evidence_str)

        for c_idx in range(1, 8):
            cell = ws_issues.cell(row=row_idx, column=c_idx)
            cell.font = regular_font
            cell.border = thin_border

    # -------------------------------------------------------------------------
    # Sheet 4: Crawled Pages
    # -------------------------------------------------------------------------
    ws_pages = wb.create_sheet(title="Crawled Pages")
    ws_pages.views.sheetView[0].showGridLines = True
    ws_pages.freeze_panes = "A2"

    page_headers = [
        "URL", "Status Code", "Title", "Title Length", "Meta Description",
        "Meta Desc Length", "H1 Tag", "Word Count", "Canonical URL",
        "Robots Directive", "Depth", "Response Time (ms)", "Content Type"
    ]
    for col_idx, h in enumerate(page_headers, start=1):
        cell = ws_pages.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center" if col_idx in (2, 4, 6, 8, 11, 12) else "left")

    pages_list = state.get("pages", [])
    for row_idx, p in enumerate(pages_list, start=2):
        status_code = p.get("status_code", 0)
        title = p.get("title", "") or ""
        meta_desc = p.get("meta_description", "") or ""

        ws_pages.cell(row=row_idx, column=1, value=p.get("url", ""))
        
        sc_cell = ws_pages.cell(row=row_idx, column=2, value=status_code)
        sc_cell.alignment = Alignment(horizontal="center")
        if 200 <= status_code < 300:
            sc_cell.fill = fill_pass
        elif 300 <= status_code < 400:
            sc_cell.fill = fill_info
        elif 400 <= status_code < 500:
            sc_cell.fill = fill_warn
        else:
            sc_cell.fill = fill_fail

        ws_pages.cell(row=row_idx, column=3, value=title)
        ws_pages.cell(row=row_idx, column=4, value=len(title)).alignment = Alignment(horizontal="center")
        ws_pages.cell(row=row_idx, column=5, value=meta_desc)
        ws_pages.cell(row=row_idx, column=6, value=len(meta_desc)).alignment = Alignment(horizontal="center")
        ws_pages.cell(row=row_idx, column=7, value=p.get("h1", "") or "")
        ws_pages.cell(row=row_idx, column=8, value=p.get("word_count", 0)).alignment = Alignment(horizontal="center")
        ws_pages.cell(row=row_idx, column=9, value=p.get("canonical", "") or "")
        ws_pages.cell(row=row_idx, column=10, value=p.get("robots", "") or "")
        ws_pages.cell(row=row_idx, column=11, value=p.get("depth", 0)).alignment = Alignment(horizontal="center")
        ws_pages.cell(row=row_idx, column=12, value=p.get("response_time_ms", 0)).alignment = Alignment(horizontal="center")
        ws_pages.cell(row=row_idx, column=13, value=p.get("content_type", ""))

        for c_idx in range(1, 14):
            cell = ws_pages.cell(row=row_idx, column=c_idx)
            cell.font = regular_font
            cell.border = thin_border

    # -------------------------------------------------------------------------
    # Sheet 5: Link Graph
    # -------------------------------------------------------------------------
    ws_links = wb.create_sheet(title="Link Graph")
    ws_links.views.sheetView[0].showGridLines = True
    ws_links.freeze_panes = "A2"

    link_headers = ["Source URL", "Target URL", "Anchor Text", "Internal/External", "Target Status", "Target Domain", "Placement"]
    for col_idx, h in enumerate(link_headers, start=1):
        cell = ws_links.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center" if col_idx in (4, 5, 7) else "left")

    links_list = state.get("links", [])
    for row_idx, l in enumerate(links_list, start=2):
        is_int = l.get("internal", True)
        ws_links.cell(row=row_idx, column=1, value=l.get("source_url", ""))
        ws_links.cell(row=row_idx, column=2, value=l.get("target_url", ""))
        ws_links.cell(row=row_idx, column=3, value=l.get("anchor_text", "") or "")
        ws_links.cell(row=row_idx, column=4, value="Internal" if is_int else "External").alignment = Alignment(horizontal="center")
        
        t_status = l.get("status_code")
        st_cell = ws_links.cell(row=row_idx, column=5, value=t_status if t_status is not None else "-")
        st_cell.alignment = Alignment(horizontal="center")
        if t_status and isinstance(t_status, int) and t_status >= 400:
            st_cell.fill = fill_fail

        ws_links.cell(row=row_idx, column=6, value=l.get("target_domain", ""))
        ws_links.cell(row=row_idx, column=7, value=l.get("placement", "body")).alignment = Alignment(horizontal="center")

        for c_idx in range(1, 8):
            cell = ws_links.cell(row=row_idx, column=c_idx)
            cell.font = regular_font
            cell.border = thin_border

    # -------------------------------------------------------------------------
    # Sheet 6: Performance & PageSpeed
    # -------------------------------------------------------------------------
    ws_perf = wb.create_sheet(title="Performance & Speed")
    ws_perf.views.sheetView[0].showGridLines = True
    ws_perf.freeze_panes = "A2"

    perf_headers = ["URL", "Status Code", "Server Response Time (ms)", "JS Render Time (ms)", "Mobile PageSpeed Score", "Desktop PageSpeed Score"]
    for col_idx, h in enumerate(perf_headers, start=1):
        cell = ws_perf.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center" if col_idx in (2, 3, 4, 5, 6) else "left")

    pagespeed_lookup = {ps.get("url"): ps for ps in state.get("pagespeed", []) if ps.get("url")}
    for row_idx, p in enumerate(pages_list, start=2):
        p_url = p.get("url", "")
        resp_time = p.get("response_time_ms", 0)
        ps_info = pagespeed_lookup.get(p_url, {})
        m_score = ps_info.get("mobile", {}).get("performance_score", "-")
        d_score = ps_info.get("desktop", {}).get("performance_score", "-")

        ws_perf.cell(row=row_idx, column=1, value=p_url)
        ws_perf.cell(row=row_idx, column=2, value=p.get("status_code", 200)).alignment = Alignment(horizontal="center")
        
        rt_cell = ws_perf.cell(row=row_idx, column=3, value=resp_time)
        rt_cell.alignment = Alignment(horizontal="center")
        if resp_time > 1500:
            rt_cell.fill = fill_fail
        elif resp_time > 500:
            rt_cell.fill = fill_warn

        ws_perf.cell(row=row_idx, column=4, value=p.get("render_time_ms") or "-").alignment = Alignment(horizontal="center")
        ws_perf.cell(row=row_idx, column=5, value=m_score).alignment = Alignment(horizontal="center")
        ws_perf.cell(row=row_idx, column=6, value=d_score).alignment = Alignment(horizontal="center")

        for c_idx in range(1, 7):
            cell = ws_perf.cell(row=row_idx, column=c_idx)
            cell.font = regular_font
            cell.border = thin_border

    # Auto-adjust column widths across all sheets
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                # Limit length computation to first 200 rows for speed
                if cell.row > 200:
                    break
                val = str(cell.value or '')
                if val:
                    max_len = max(max_len, len(val))
            sheet.column_dimensions[col_letter].width = min(60, max(max_len + 3, 12))

    wb.save(filepath)
    return True
